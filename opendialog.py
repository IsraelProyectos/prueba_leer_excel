
import wx

from openpyxl import *
from openpyxl.styles import Color, PatternFill, Font, Border

class MyFrame(wx.Frame):
            def __init__(self, parent, title):
                  wx.Frame.__init__(self, parent, title=title, size=(500,200))

                  self.pathFile = ''
                  self.txtRuta = wx.TextCtrl(self, pos=(80,50), size=(150,20), style=wx.TE_READONLY)
                  self.buttonFind = wx.Button(self, label="Buscar...", pos=(260,50), size=(100,20))
                  self.buttonFind.Bind(wx.EVT_BUTTON, self.openFile)
                  self.buttonExecute = wx.Button(self, label="Convertir", pos=(260,100), size=(100,20))
                  self.buttonExecute.Bind(wx.EVT_BUTTON, self.createExcel)
                  self.buttonExecute.Disable()
                  self.labelEstadoOperacion= wx.StaticText(self, pos=(80,130), size=(280,20), style=wx.TE_READONLY)
                  self.labelEstadoOperacion.SetForegroundColour( wx.Colour( 0, 138, 0))

                  self.columna_excel = [ ]
                  self.todas_columnas = [ ]
                  self.registro_excel_final = [ ]
                  self.registros_excel_final =[ ]
                  self.email = ""
                  self.i = 0
                  self.z = 1
                  self.Show(True)

            def createExcel(self, e):

				try:
					#Cargando fichero desde textBox, obtenido de openFileDialog
					doc = load_workbook(self.pathFile)
					hoja = doc.worksheets[0]
					#hoja = doc.get_sheet_names()
					print(hoja)


					#Leyendo filas del excel y guardandola en una lista
					for fila in hoja.rows:
						for columna in fila:
							self.columna_excel.append(columna.value)

						#Guardando la lista dentro de otra lista para tener las filas separadas
						# if self.i !=0:
						self.todas_columnas.append(self.columna_excel[0:-1])
						self.columna_excel = [ ]
						self.i +=1

					#Juntado los registros por email
					for fila in self.todas_columnas:
							fila = [fila[2], fila[1], fila[0], fila[3], fila[4], fila[5], fila[6]]
							#La primera comparacion siempre sera nula e ira al else
							if self.email == fila[0]:
								#Anadiendo cod_instalacion, objetivoAO y objetivoAOA al registro con el mismo mail
								self.registros_excel_final[-1][2] = self.registros_excel_final[-1][2] + "; " + fila[2]
								self.registros_excel_final[-1][3] = str(self.registros_excel_final[-1][3]) + "; " + str(fila[3])
								self.registros_excel_final[-1][4] = str(self.registros_excel_final[-1][4]) + "; " + str(fila[4])
							else:
								#Anadiendo fila nueva
								self.registro_excel_final.append(fila)
								self.registros_excel_final.append(self.registro_excel_final[-1])
								self.registro_excel_final = [ ]

							#Guardando el mail de la fila insertada anteriormente para aplicar la comparacion
							self.email = fila[0]

					#Creando el excel de salida
					book = Workbook()
					hoja1 = book.active

					#Recorriendo los registros con el mismo mail y insertandolos en el Excel creado anteriormente
					for regs in self.registros_excel_final:
							y=1
							for reg in regs:
								#print(reg)
								celda = hoja1.cell(row=self.z, column=y).value = reg
								if self.z == 1:
									x=1
									for celda in reg:
										greyFill = PatternFill(start_color='A9A9A9', end_color='A9A9A9', fill_type='solid')
										hoja1.cell(row=1, column=x).fill = greyFill
										x+=1
								y+=1
							self.z+=1
					#Guardando el WorkBook donde seleccione el Usuario
					with wx.FileDialog(self, "Save XLSX file", wildcard="XLSX files (*.xlsx)|*.xlsx",
					   style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:
					  if fileDialog.ShowModal() == wx.ID_CANCEL:
					  	return
					  #Guardando en variable el path de donde se quiere guardar el archivo
					  pathname = fileDialog.GetPath()
					  try:
					  	with open(pathname, 'w') as file:
					  		book.save(pathname)
					  except IOError:
					  	wx.LogError("Cannot save current data in file '%s'." % pathname)
					  self.labelEstadoOperacion.SetLabel("El fichero se ha creado correctamente")
				except KeyError:
					self.labelEstadoOperacion.SetForegroundColour( wx.Colour( 255, 0, 0))
					self.labelEstadoOperacion.SetLabel("No se ha podido crear el fichero")
            
            def openFile(self, e):
				try:
	            	#Creando la ventana para escoger el archivo. Solo para archivos con extension .xlsx(Archivo excel 2010)
					with wx.FileDialog(self, "Open XLSX file", wildcard="XLSX files (*.xlsx)|*.xlsx",
						style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:

						#Creando opciones de la ventana(Abrir o Cancelar)
						if fileDialog.ShowModal() == wx.ID_CANCEL:
							return

				        #Guardando el path del archivo en variable
				        self.pathFile = fileDialog.GetPath()

				        #Asignando ese path al textBox
				        try:
				            self.txtRuta.SetValue(self.pathFile)
				            self.buttonExecute.Enable()
				        except IOError:
							wx.LogError("Cannot open file '%s'." % newfile)
				except Keyerror:
					print(err)
app = wx.App(False)
frame = MyFrame(None, 'Creacion Fichero Excel')
app.MainLoop()
