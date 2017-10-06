#!/usr/bin/env python
#coding=utf-8

import wx
from openpyxl import *
from openpyxl.styles import Color, PatternFill, Font, Border, colors, borders, Side

class MyFrame(wx.Frame):
            def __init__(self, parent, title):
                  wx.Frame.__init__(self, parent, title=title, size=(400,200))

                  self.pathFile = ''
                  self.txtRuta = wx.TextCtrl(self, pos=(10,50), size=(250,20), style=wx.TE_READONLY)
                  self.buttonFind = wx.Button(self, label="Buscar...", pos=(270,50), size=(100,20))
                  self.buttonFind.Bind(wx.EVT_BUTTON, self.openFile)
                  self.buttonExecute = wx.Button(self, label="Convertir", pos=(270,100), size=(100,20))
                  self.buttonExecute.Bind(wx.EVT_BUTTON, self.createExcel)
                  self.buttonExecute.Disable()
                  self.labelEstadoOperacion= wx.StaticText(self, pos=(10,130), size=(360,20), style=wx.TE_READONLY)
                  # self.labelEstadoOperacion.SetBackgroundColour( wx.Colour( 255, 255, 255))
                  self.labelEstadoOperacion.SetForegroundColour(wx.Colour( 0, 138, 0))
                  

                  self.columna_excel = [ ]
                  self.todas_columnas = [ ]
                  self.registro_excel_final = [ ]
                  self.registros_excel_final =[ ]
                  self.fields = ['EMAIL', 'CODIGO_1', 'CODIGO_2', 'CODIGO_3', 'CODIGO_4', 'CODIGO_5', 'NOMBRE_CODIGO_1',
                  						  'NOMBRE_CODIGO_2', 'NOMBRE_CODIGO_3', 'NOMBRE_CODIGO_4', 'NOMBRE_CODIGO_5', 	
										  'OBJ_AO_1', 'OBJ_AO_2', 'OBJ_AO_3', 'OBJ_AO_4', 'OBJ_AO_5', 
										  'OBJ_AOA_1', 'OBJ_AOA_2',	'OBJ_AOA_3', 'OBJ_AOA_4', 'OBJ_AOA_5',	
										  'EMAIL_CC', 'EMAIL_REMITENTE', 'EMAIL_CONTACTO', 'NOMBRE']
                  self.email = ""
                  self.i = 0
                  self.z = 2
                  self.columnaDelExcel = 1
                  self.Centre(True)
                  self.SetBackgroundColour(wx.Colour( 252, 255, 228))
                  self.Show(True)

            def createExcel(self, e):

				try:
					#Cargando fichero desde textBox, obtenido de openFileDialog
					doc = load_workbook(self.pathFile)
					hoja = doc.worksheets[0]
					#print(doc.sheet_names())
					
					#Leyendo filas del excel y guardandola en una lista
					i=0
					for fila in hoja.rows:
						
						if i != 0:

							for columna in fila:
								self.columna_excel.append(columna.value)

							#Campos agregados de cod_instalacion
							self.columna_excel.insert(1,'')
							self.columna_excel.insert(2,'')
							self.columna_excel.insert(3,'')
							self.columna_excel.insert(4,'')

							#Campos agregados de concesionario
							self.columna_excel.insert(9,'')
							self.columna_excel.insert(10,'')
							self.columna_excel.insert(11,'')
							self.columna_excel.insert(12,'')

							#Campos agregados de objectivo_AOA
							self.columna_excel.insert(14,'')
							self.columna_excel.insert(15,'')
							self.columna_excel.insert(16,'')
							self.columna_excel.insert(17,'')

							#Campos agregados de objectivo_AO
							self.columna_excel.insert(6,'')
							self.columna_excel.insert(7,'')
							self.columna_excel.insert(8,'')
							self.columna_excel.insert(9,'')

							
							
							

							#Guardando la lista dentro de otra lista para tener las filas separadas
							#print(self.columna_excel[8])
							self.todas_columnas.append(self.columna_excel)
							self.columna_excel = [ ]
							self.i +=1
						
						i=i+1
					# print(self.todas_columnas[0])
					#print(self.todas_columnas)
					email='hola'
					i=-1
					y=1
					x=13
					w=18
					t=6
					nombreImpacto=''

					for registro in self.todas_columnas:
						
						if email == registro[10]:
							# print(registro[0])
							# print(self.registros_excel_final[i][0])
							
							self.registros_excel_final[i][y] = registro[0]
							self.registros_excel_final[i][x] = registro[12]
							self.registros_excel_final[i][w] = registro[17]
							self.registros_excel_final[i][t] = registro[5]

							# if registro[7].title() != nombreImpacto.title():
							# 	self.registros_excel_final[i][7] = ''
							
							x=x+1
							y=y+1
							w=w+1
							t=t+1

						else:
							#print(registro)
							
							email=registro[10]
							#print(email)
							if registro[11] is not None:
								registro[11] = registro[11].title()
								nombreImpacto=registro[11].title()
							self.registros_excel_final.append(registro)
							# print(nombreImpacto)
							i=i+1
							y=1
							x=13
							w=18
							t=6
						#print(self.registros_excel_final[0])
						#print(self.registros_excel_final)

					# print(self.registros_excel_final[0])
					#print(i)

					#Creando el excel de salida
					book = Workbook()
					hoja1 = book.active
					
					

					# for countRegistro in self.registros_excel_final:
					# 	for countReg in countRegistro:
					# 		print(len(countRegistro[2]))

					#Recorriendo los registros con el mismo mail y insertandolos en el Excel creado anteriormente
					insercionExcel = 0
					#print(self.registros_excel_final[0])
					for regs in self.registros_excel_final:

							if insercionExcel != 0:
								if insercionExcel != 0:
									regsInverse = [
									regs[10],
									regs[0],
									regs[1],
									regs[2],
									regs[3],
									regs[4],
									regs[5],
									regs[6],
									regs[7],
									regs[8],
									regs[9],
									regs[17],
									regs[18],
									regs[19],
									regs[20],
									regs[21],
									regs[12],
									regs[13],
									regs[14],
									regs[15],
									regs[16],
									regs[22],
									regs[23],
									regs[24],
									regs[11]]
								y=1
								for reg in regsInverse:
									celda = hoja1.cell(row=self.z, column=y).value = reg
									y+=1
								self.z+=1
							else:
								sig=1
								for i in [self.fields[0], self.fields[1], self.fields[2], self.fields[3], self.fields[4], self.fields[5], self.fields[6],	
										  self.fields[7], self.fields[8], self.fields[9], self.fields[10], self.fields[11], 
										  self.fields[12], self.fields[13], self.fields[14], self.fields[15], self.fields[16],	
										  self.fields[17], self.fields[18], self.fields[19], self.fields[20], self.fields[21],
										  self.fields[22], self.fields[23], self.fields[24]]:
										  celda = hoja1.cell(row=1, column=sig).value = i
										  sig=sig+1
								insercionExcel+=1

					#print(self.registros_excel_final)
					#Guardando el WorkBook donde seleccione el Usuario
					with wx.FileDialog(self, "Save XLSX file", wildcard="XLSX files (*.xlsx)|*.xlsx",
					   style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT) as fileDialog:
					  if fileDialog.ShowModal() == wx.ID_CANCEL:
					  	return
					  #Guardando en variable el path de donde se quiere guardar el archivo
					  pathname = fileDialog.GetPath()
					  try:
					  	with open(pathname, 'w') as file:
					  		colorFill = PatternFill(start_color='A8A8A8', end_color='A8A8A8', fill_type='solid')
							font = Font(color=colors.BLACK, italic=True, bold=True)
							border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
							# hoja1['A1':'Y1'].fill = redFill
							# hoja1['A1':'Y1'].font = Font(color=colors.BLACK, italic=True, bold=True)
							for cell in hoja1["1:1"]:
								cell.font = font
								cell.fill = colorFill
								cell.border = border
					  		book.save(pathname)
					  except IOError:
					  	wx.LogError("Cannot save current data in file '%s'." % pathname)
					  self.labelEstadoOperacion.SetForegroundColour(wx.Colour( 0, 138, 0))
					  self.labelEstadoOperacion.SetLabel("El fichero se ha creado correctamente")
				except KeyError:
					self.labelEstadoOperacion.SetForegroundColour(wx.Colour(255, 0, 0))
					self.labelEstadoOperacion.SetLabel("No se ha podido crear el fichero")
				except IndexError:
					self.labelEstadoOperacion.SetForegroundColour(wx.Colour(255, 0, 0))
					self.labelEstadoOperacion.SetLabel("El formato de celdas del documento no es válido")

            
            def openFile(self, e):
				try:
					self.labelEstadoOperacion.SetLabel("")

	            	#Creando la ventana para escoger el archivo. Solo para archivos con extension .xlsx(Archivo excel 2010)
					with wx.FileDialog(self, "Abrir archivo .xlsx", wildcard="XLSX files (*.xlsx)|*.xlsx",
						style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:

						#Cerrar ventana de diálogo al dar a cancelar
						if fileDialog.ShowModal() == wx.ID_CANCEL:
							return

				        #Guardando el path del archivo en variable
				        self.pathFile = fileDialog.GetPath()

				        #Asignando ese path al textBox
				        try:
				            self.txtRuta.SetValue(self.pathFile)
				            self.buttonExecute.Enable()
				        except IOError:
							wx.LogError("Cannot open file '%s'." % newfile)
				except Keyerror:
					print(err)
app = wx.App(False)
frame = MyFrame(None, 'Creación Fichero Excel')
app.MainLoop()
